import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import PyPDF2
import re
import sqlite3
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# --- CONFIGURAZIONE ---
st.set_page_config(page_title="Programma Canile", layout="centered")

def init_db():
    """
    Inizializza il database creando le tabelle necessarie.
    NOVITÀ: Aggiunta tabella 'programmi' per salvare i programmi approvati
    """
    conn = sqlite3.connect('canile.db')
    c = conn.cursor()
    
    # Tabella storico turni (già esistente)
    c.execute('CREATE TABLE IF NOT EXISTS storico (data TEXT, inizio TEXT, cane TEXT, volontario TEXT, luogo TEXT)')
    
    # Tabella anagrafica cani (già esistente)
    c.execute('''CREATE TABLE IF NOT EXISTS anagrafica_cani 
                 (nome TEXT PRIMARY KEY, cibo TEXT, guinzaglieria TEXT, strumenti TEXT, 
                  attivita TEXT, note TEXT, tempo TEXT)''')
    
    # NUOVA TABELLA: Programmi salvati
    c.execute('''CREATE TABLE IF NOT EXISTS programmi 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  data TEXT NOT NULL,
                  nome TEXT,
                  creato_il TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  turni_json TEXT NOT NULL,
                  note TEXT,
                  UNIQUE(data))''')
    
    # NUOVA TABELLA: Backup automatico dello storico
    c.execute('''CREATE TABLE IF NOT EXISTS storico_backup 
                 (backup_id INTEGER PRIMARY KEY AUTOINCREMENT,
                  backup_timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  data TEXT, inizio TEXT, cane TEXT, volontario TEXT, luogo TEXT)''')
    
    conn.commit()
    conn.close()

def backup_storico():
    """
    Crea un backup automatico dello storico prima di modifiche importanti.
    Questo garantisce che lo storico non vada mai perso.
    """
    conn = sqlite3.connect('canile.db')
    c = conn.cursor()
    
    # Copio tutti i dati da storico a storico_backup
    c.execute('''INSERT INTO storico_backup (data, inizio, cane, volontario, luogo)
                 SELECT data, inizio, cane, volontario, luogo FROM storico''')
    
    conn.commit()
    
    # Conto i backup
    count = c.execute("SELECT COUNT(DISTINCT backup_timestamp) FROM storico_backup").fetchone()[0]
    conn.close()
    
    return count

def ripristina_storico_da_backup(backup_timestamp=None):
    """
    Ripristina lo storico da un backup specifico o dall'ultimo disponibile.
    """
    conn = sqlite3.connect('canile.db')
    c = conn.cursor()
    
    if backup_timestamp:
        # Ripristino da un backup specifico
        c.execute("DELETE FROM storico")
        c.execute('''INSERT INTO storico (data, inizio, cane, volontario, luogo)
                     SELECT data, inizio, cane, volontario, luogo 
                     FROM storico_backup 
                     WHERE backup_timestamp = ?''', (backup_timestamp,))
    else:
        # Ripristino dall'ultimo backup
        ultimo_backup = c.execute("SELECT MAX(backup_timestamp) FROM storico_backup").fetchone()[0]
        if ultimo_backup:
            c.execute("DELETE FROM storico")
            c.execute('''INSERT INTO storico (data, inizio, cane, volontario, luogo)
                         SELECT data, inizio, cane, volontario, luogo 
                         FROM storico_backup 
                         WHERE backup_timestamp = ?''', (ultimo_backup,))
    
    conn.commit()
    conn.close()

def salva_programma(data, turni_list, nome=None, note=None):
    """
    Salva un programma approvato nel database.
    Se esiste già un programma per quella data, lo sovrascrive.
    
    Args:
        data: data del programma (formato YYYY-MM-DD)
        turni_list: lista di dizionari con i turni
        nome: nome descrittivo del programma (opzionale)
        note: note aggiuntive (opzionale)
    
    Returns:
        True se salvato con successo, False altrimenti
    """
    conn = sqlite3.connect('canile.db')
    c = conn.cursor()
    
    try:
        # Converto la lista di turni in JSON
        turni_json = json.dumps(turni_list, ensure_ascii=False)
        
        # Inserisco o aggiorno il programma
        c.execute('''INSERT OR REPLACE INTO programmi (data, nome, turni_json, note, creato_il)
                     VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)''',
                  (data, nome, turni_json, note))
        
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        st.error(f"Errore nel salvataggio: {str(e)}")
        conn.close()
        return False

def carica_programma(data):
    """
    Carica un programma salvato dal database.
    
    Args:
        data: data del programma (formato YYYY-MM-DD)
    
    Returns:
        Dizionario con i dati del programma o None se non trovato
    """
    conn = sqlite3.connect('canile.db')
    c = conn.cursor()
    
    row = c.execute('''SELECT id, data, nome, creato_il, turni_json, note 
                       FROM programmi WHERE data = ?''', (data,)).fetchone()
    
    conn.close()
    
    if row:
        return {
            'id': row[0],
            'data': row[1],
            'nome': row[2],
            'creato_il': row[3],
            'turni': json.loads(row[4]),
            'note': row[5]
        }
    return None

def elimina_programma(data):
    """
    Elimina un programma salvato.
    
    Args:
        data: data del programma da eliminare
    
    Returns:
        True se eliminato con successo
    """
    conn = sqlite3.connect('canile.db')
    c = conn.cursor()
    
    c.execute("DELETE FROM programmi WHERE data = ?", (data,))
    conn.commit()
    conn.close()
    
    return True

def lista_programmi_salvati():
    """
    Restituisce la lista di tutti i programmi salvati.
    
    Returns:
        DataFrame con i programmi salvati
    """
    conn = sqlite3.connect('canile.db')
    df = pd.read_sql_query('''SELECT id, data, nome, creato_il, 
                              LENGTH(turni_json) as dim_dati, note 
                              FROM programmi 
                              ORDER BY data DESC''', conn)
    conn.close()
    
    return df

def export_programma_excel(turni_list, data_str, nome_file=None):
    """
    Esporta un programma in formato Excel con formattazione professionale.
    
    Args:
        turni_list: lista di turni
        data_str: data del programma
        nome_file: nome del file (opzionale)
    
    Returns:
        BytesIO object con il file Excel
    """
    if nome_file is None:
        nome_file = f"Programma_{data_str}.xlsx"
    
    # Creo il workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Programma Giornaliero"
    
    # Stili
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titolo
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = f"PROGRAMMA GIORNALIERO - {data_str}"
    title_cell.font = Font(size=16, bold=True, color="4472C4")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Intestazioni colonne
    headers = ['Orario', 'Cane', 'Volontario', 'Luogo', 'Note']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dati
    for row_idx, turno in enumerate(turni_list, 4):
        ws.cell(row=row_idx, column=1, value=turno.get('Orario', ''))
        ws.cell(row=row_idx, column=2, value=turno.get('Cane', ''))
        ws.cell(row=row_idx, column=3, value=turno.get('Volontario', ''))
        ws.cell(row=row_idx, column=4, value=turno.get('Luogo', ''))
        ws.cell(row=row_idx, column=5, value=turno.get('Note', ''))
        
        # Applico i bordi
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=col).alignment = Alignment(vertical='center')
    
    # Dimensioni colonne
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    
    # Salvo in BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def export_programma_pdf(turni_list, data_str, nome_file=None):
    """
    Esporta un programma in formato PDF con formattazione professionale.
    
    Args:
        turni_list: lista di turni
        data_str: data del programma
        nome_file: nome del file (opzionale)
    
    Returns:
        BytesIO object con il file PDF
    """
    if nome_file is None:
        nome_file = f"Programma_{data_str}.pdf"
    
    # Creo il buffer
    buffer = io.BytesIO()
    
    # Creo il documento
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Stili
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#4472C4'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Titolo
    title = Paragraph(f"<b>PROGRAMMA GIORNALIERO - {data_str}</b>", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    # Preparo i dati per la tabella
    data = [['Orario', 'Cane', 'Volontario', 'Luogo', 'Note']]
    
    for turno in turni_list:
        data.append([
            turno.get('Orario', ''),
            turno.get('Cane', ''),
            turno.get('Volontario', ''),
            turno.get('Luogo', ''),
            turno.get('Note', '')[:30]  # Limito le note per evitare overflow
        ])
    
    # Creo la tabella
    table = Table(data, colWidths=[3*cm, 4*cm, 4.5*cm, 4*cm, 4*cm])
    
    # Stile della tabella
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Corpo
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
    ]))
    
    elements.append(table)
    
    # Genero il PDF
    doc.build(elements)
    
    buffer.seek(0)
    return buffer

def load_gsheets(sheet_name):
    """
    Carica i dati da Google Sheets con gestione errori dettagliata.
    Restituisce: (DataFrame, messaggio_errore)
    """
    # Link al tuo Google Sheet (assicurati che sia pubblico o accessibile)
    url = f"https://docs.google.com/spreadsheets/d/1pcFa454IT1tlykbcK-BeAU9hnIQ_D8V_UuZaKI_KtYM/gviz/tq?tqx=out:csv&sheet={sheet_name}"
    try:
        df = pd.read_csv(url)
        df.columns = [c.strip().lower() for c in df.columns]
        
        # ### GESTIONE COLONNA 'automatico' per Luoghi
        if sheet_name == "Luoghi" and 'automatico' not in df.columns:
            df['automatico'] = 'sì'
        
        # ### GESTIONE COLONNA 'adiacente' per Luoghi
        if sheet_name == "Luoghi" and 'adiacente' not in df.columns:
            df['adiacente'] = ''
        
        # ### GESTIONE COLONNA 'reattività' per Cani
        if sheet_name == "Cani" and 'reattività' not in df.columns:
            df['reattività'] = 0
        elif sheet_name == "Cani":
            # Converto a numerico, mettendo 0 dove non valido
            df['reattività'] = pd.to_numeric(df['reattività'], errors='coerce').fillna(0)
        
        df_clean = df.dropna(how='all')
        return df_clean, None
    except Exception as e:
        return pd.DataFrame(), str(e)

def parse_pdf_content(text):
    """
    Estrae i campi dal PDF cercando i titoli in MAIUSCOLO.
    I titoli sono: CIBO, GUINZAGLIERIA, STRUMENTI, ATTIVITÀ, NOTE, TEMPO
    Il contenuto di ogni campo è il testo che segue il titolo fino al prossimo titolo o fino alla fine.
    """
    # Lista dei campi da estrarre (nell'ordine in cui appaiono nel PDF)
    campi = ['CIBO', 'GUINZAGLIERIA', 'STRUMENTI', 'ATTIVITÀ', 'NOTE', 'TEMPO']
    dati_estratti = {c: "N/D" for c in campi}
    
    # Pulizia preliminare del testo
    text = text.replace('\n\n', '\n').replace('\r', '')
    
    for i, campo in enumerate(campi):
        # Pattern migliorato: cerca il campo in maiuscolo (con possibili : o spazi dopo)
        # e cattura tutto fino al prossimo campo maiuscolo o fine testo
        
        # Creo il pattern per il campo successivo (se esiste)
        if i < len(campi) - 1:
            # Non è l'ultimo campo: cerco fino al prossimo campo
            prossimi_campi = '|'.join(campi[i+1:])
            pattern = rf"{campo}[\s:]*\n+(.*?)(?=\n+(?:{prossimi_campi})[\s:]|\Z)"
        else:
            # È l'ultimo campo (TEMPO): cerco fino alla fine
            pattern = rf"{campo}[\s:]*\n+(.*?)(?=\Z)"
        
        match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
        if match:
            contenuto = match.group(1).strip()
            # Rimuovo eventuali righe vuote multiple
            contenuto = re.sub(r'\n\s*\n', '\n', contenuto)
            dati_estratti[campo] = contenuto if contenuto else "N/D"
        else:
            # Tentativo alternativo: cerca il campo seguito da qualsiasi testo fino al prossimo campo in maiuscolo
            pattern_alt = rf"{campo}[\s:]*(.+?)(?=(?:{'|'.join(campi[i+1:]) if i < len(campi)-1 else 'XXXXXX'})[\s:]|\Z)"
            match_alt = re.search(pattern_alt, text, re.DOTALL | re.IGNORECASE)
            if match_alt:
                contenuto = match_alt.group(1).strip()
                contenuto = re.sub(r'\n\s*\n', '\n', contenuto)
                dati_estratti[campo] = contenuto if contenuto else "N/D"
    
    return dati_estratti

def get_reattivita_cane(nome_cane, df_cani):
    """Restituisce il valore di reattività di un cane dal DataFrame"""
    if df_cani.empty or 'reattività' not in df_cani.columns:
        return 0
    riga = df_cani[df_cani['nome'] == nome_cane]
    if not riga.empty:
        return float(riga.iloc[0]['reattività'])
    return 0

def get_campi_adiacenti(campo, df_luoghi):
    """
    Restituisce la lista dei campi adiacenti a un dato campo leggendo dal DataFrame Luoghi.
    La colonna 'adiacente' può contenere nomi separati da virgola, es: "Campo1, Campo2"
    """
    if df_luoghi.empty or 'adiacente' not in df_luoghi.columns:
        return []
    
    riga = df_luoghi[df_luoghi['nome'] == campo]
    if not riga.empty:
        adiacenti_str = str(riga.iloc[0]['adiacente']).strip()
        if adiacenti_str and adiacenti_str != 'nan':
            # Separo per virgola e pulisco gli spazi
            return [c.strip() for c in adiacenti_str.split(',') if c.strip()]
    return []

def campo_valido_per_reattivita(cane, campo, turni_attuali, ora_attuale_str, df_cani, df_luoghi):
    """
    Verifica se un campo è valido per un cane considerando la reattività.
    CONTROLLO BIDIREZIONALE:
    - Se il cane DA ASSEGNARE ha reattività > 5, verifica che nei campi adiacenti non ci siano altri cani
    - Se nei campi adiacenti ci sono CANI CON REATTIVITÀ > 5, il campo non è valido
    
    Args:
        cane: nome del cane da verificare
        campo: nome del campo da verificare
        turni_attuali: lista di tutti i turni già programmati (automatici + manuali)
        ora_attuale_str: orario del turno da verificare (formato "HH:MM")
        df_cani: DataFrame con i dati dei cani (include colonna reattività)
        df_luoghi: DataFrame con i dati dei luoghi (include colonna adiacente)
    
    Returns:
        True se il campo è valido, False se ci sono conflitti di reattività
    """
    reattivita_cane_corrente = get_reattivita_cane(cane, df_cani)
    campi_adiacenti = get_campi_adiacenti(campo, df_luoghi)
    
    # Verifico i cani già presenti nei campi adiacenti allo stesso orario
    for turno in turni_attuali:
        if turno["Orario"] == ora_attuale_str:
            if turno["Luogo"] in campi_adiacenti:
                # C'è un cane in un campo adiacente
                cane_adiacente = turno["Cane"]
                
                # Ignoro i turni speciali (Briefing, Pasti)
                if cane_adiacente in ["TUTTI", "Da assegnare"]:
                    continue
                
                reattivita_cane_adiacente = get_reattivita_cane(cane_adiacente, df_cani)
                
                # CONFLITTO se ALMENO UNO dei due ha reattività > 5
                if reattivita_cane_corrente > 5 or reattivita_cane_adiacente > 5:
                    return False
    
    return True

def get_info_cane(nome_cane):
    """
    Recupera le informazioni complete di un cane dall'anagrafica.
    Restituisce un dizionario con tutti i campi, o valori "N/D" se il cane non è trovato.
    """
    conn = sqlite3.connect('canile.db')
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute("SELECT * FROM anagrafica_cani WHERE nome=?", (nome_cane.capitalize(),)).fetchone()
        if row:
            return {
                'CIBO': row['cibo'] or 'N/D',
                'GUINZAGLIERIA': row['guinzaglieria'] or 'N/D',
                'STRUMENTI': row['strumenti'] or 'N/D',
                'ATTIVITÀ': row['attivita'] or 'N/D',
                'NOTE': row['note'] or 'N/D',
                'TEMPO': row['tempo'] or 'N/D'
            }
        else:
            return {k: "N/D" for k in ['CIBO', 'GUINZAGLIERIA', 'STRUMENTI', 'ATTIVITÀ', 'NOTE', 'TEMPO']}
    finally:
        conn.close()

def salva_info_cane(nome_cane, dati):
    """
    Salva o aggiorna le informazioni di un cane nell'anagrafica SQLite.
    
    Args:
        nome_cane: nome del cane (verrà capitalizzato)
        dati: dizionario con le chiavi CIBO, GUINZAGLIERIA, STRUMENTI, ATTIVITÀ, NOTE, TEMPO
    
    Returns:
        True se salvato con successo, False altrimenti
    """
    conn = sqlite3.connect('canile.db')
    try:
        conn.execute('''INSERT OR REPLACE INTO anagrafica_cani 
                        (nome, cibo, guinzaglieria, strumenti, attivita, note, tempo)
                        VALUES (?, ?, ?, ?, ?, ?, ?)''',
                     (nome_cane.capitalize(),
                      dati.get('CIBO', 'N/D'),
                      dati.get('GUINZAGLIERIA', 'N/D'),
                      dati.get('STRUMENTI', 'N/D'),
                      dati.get('ATTIVITÀ', 'N/D'),
                      dati.get('NOTE', 'N/D'),
                      dati.get('TEMPO', 'N/D')))
        conn.commit()
        return True
    except Exception as e:
        st.error(f"Errore nel salvataggio in anagrafica: {str(e)}")
        return False
    finally:
        conn.close()

def get_esperienza_volontari(cane, df_storico):
    """
    Calcola l'esperienza di ciascun volontario con un dato cane basandosi sullo storico.
    
    Args:
        cane: nome del cane
        df_storico: DataFrame dello storico turni
    
    Returns:
        Dizionario {volontario: numero_turni_fatti}
    """
    if df_storico.empty:
        return {}
    
    # Filtro i turni per il cane specifico
    df_cane = df_storico[df_storico['cane'] == cane]
    
    # Conto i turni per volontario
    esperienza = df_cane.groupby('volontario').size().to_dict()
    
    return esperienza

def assegna_volontario_esperto(cane, volontari_disponibili, df_storico):
    """
    Assegna il volontario più esperto con un dato cane, se disponibile.
    
    Args:
        cane: nome del cane
        volontari_disponibili: lista dei volontari disponibili
        df_storico: DataFrame dello storico turni
    
    Returns:
        Nome del volontario assegnato (il più esperto, o il primo della lista se nessuno ha esperienza)
    """
    esperienza = get_esperienza_volontari(cane, df_storico)
    
    # Filtro solo i volontari disponibili che hanno esperienza
    volontari_con_esperienza = {v: esperienza.get(v, 0) for v in volontari_disponibili if esperienza.get(v, 0) > 0}
    
    if volontari_con_esperienza:
        # Restituisco il volontario con più esperienza
        return max(volontari_con_esperienza, key=volontari_con_esperienza.get)
    else:
        # Nessuno ha esperienza: restituisco il primo disponibile
        return volontari_disponibili[0] if volontari_disponibili else "Da assegnare"

def salva_turni_in_storico(turni, data):
    """
    Salva i turni nello storico del database (solo turni con cani specifici, esclude briefing/pasti).
    
    Args:
        turni: lista di dizionari con i turni
        data: oggetto datetime della data
    
    Returns:
        Numero di record salvati
    """
    conn = sqlite3.connect('canile.db')
    c = conn.cursor()
    
    data_str = data.strftime('%Y-%m-%d')
    count = 0
    
    for turno in turni:
        cane = turno.get("Cane", "")
        # Salvo solo i turni con cani specifici (non briefing, non pasti, non "Da assegnare")
        if cane and cane not in ["TUTTI", "Da assegnare", ""]:
            c.execute('''INSERT INTO storico (data, inizio, cane, volontario, luogo)
                         VALUES (?, ?, ?, ?, ?)''',
                      (data_str,
                       turno.get("Orario", ""),
                       cane,
                       turno.get("Volontario", ""),
                       turno.get("Luogo", "")))
            count += 1
    
    conn.commit()
    conn.close()
    
    return count

# --- INIZIALIZZAZIONE ---
init_db()

# Inizializzo lo stato della sessione
if 'programma' not in st.session_state:
    st.session_state.programma = []

# --- INTERFACCIA PRINCIPALE ---
st.title("🐕 Gestione Programma Canile")

# --- CARICAMENTO DATI DA GOOGLE SHEETS CON DIAGNOSTICA ---
st.sidebar.header("📊 Stato Caricamento Dati")

# Carico i dati da Google Sheets
df_cani, err_cani = load_gsheets("Cani")
df_volontari, err_volontari = load_gsheets("Volontari")
df_luoghi, err_luoghi = load_gsheets("Luoghi")

# Mostro lo stato del caricamento
with st.sidebar:
    if not df_cani.empty:
        st.success(f"✅ Cani: {len(df_cani)} caricati")
    else:
        st.error(f"❌ Cani: Errore caricamento")
        if err_cani:
            st.caption(f"Dettagli: {err_cani}")
    
    if not df_volontari.empty:
        st.success(f"✅ Volontari: {len(df_volontari)} caricati")
    else:
        st.error(f"❌ Volontari: Errore caricamento")
        if err_volontari:
            st.caption(f"Dettagli: {err_volontari}")
    
    if not df_luoghi.empty:
        st.success(f"✅ Luoghi: {len(df_luoghi)} caricati")
    else:
        st.error(f"❌ Luoghi: Errore caricamento")
        if err_luoghi:
            st.caption(f"Dettagli: {err_luoghi}")
    
    # Pulsante di ricaricamento
    if st.button("🔄 Ricarica Dati", use_container_width=True):
        st.rerun()
    
    st.divider()
    
    # Info database
    st.caption("💾 Database locale: canile.db")
    conn_check = sqlite3.connect('canile.db')
    num_storico = pd.read_sql_query("SELECT COUNT(*) as count FROM storico", conn_check).iloc[0]['count']
    num_anagrafica = pd.read_sql_query("SELECT COUNT(*) as count FROM anagrafica_cani", conn_check).iloc[0]['count']
    conn_check.close()
    st.caption(f"📊 Storico turni: {num_storico}")
    st.caption(f"📋 Anagrafica cani: {num_anagrafica}")

# Carico lo storico per l'assegnazione intelligente
conn_storico = sqlite3.connect('canile.db')
df_storico = pd.read_sql_query("SELECT * FROM storico", conn_storico)
conn_storico.close()

# --- TAB PRINCIPALI ---
tab_programma, tab_gestione, tab_anagrafica, tab_storico = st.tabs([
    "📅 Crea Programma", 
    "💾 Gestione Programmi Salvati",
    "📋 Anagrafica Cani", 
    "📊 Storico & Statistiche"
])

# ============================================================================
# TAB 1: CREA PROGRAMMA (Codice esistente)
# ============================================================================
with tab_programma:
    st.header("Pianificazione Giornaliera")
    
    # Sezione configurazione
    st.subheader("⚙️ Configurazione")
    
    col1, col2 = st.columns(2)
    
    with col1:
        data_t = st.date_input("📅 Data del turno", datetime.today())
        ora_inizio = st.time_input("🕐 Orario inizio turni", value=datetime.strptime("08:00", "%H:%M").time())
        ora_fine = st.time_input("🕐 Orario fine turni", value=datetime.strptime("18:00", "%H:%M").time())
    
    with col2:
        durata = st.selectbox("⏱️ Durata turno (minuti)", [20, 30, 40, 60], index=1)
        num_volontari = st.number_input("👥 Numero volontari disponibili", min_value=1, max_value=20, value=4)
    
    # Briefing e Pasti
    st.subheader("🍽️ Impostazioni Briefing e Pasti")
    
    col_b1, col_b2 = st.columns(2)
    
    with col_b1:
        briefing_enabled = st.checkbox("Abilita Briefing mattutino", value=True)
        if briefing_enabled:
            ora_briefing = st.time_input("Orario Briefing", value=datetime.strptime("07:45", "%H:%M").time())
    
    with col_b2:
        pasti_enabled = st.checkbox("Abilita Pasti", value=True)
        if pasti_enabled:
            ora_pasti = st.time_input("Orario Pasti", value=datetime.strptime("13:00", "%H:%M").time())
    
    # Generazione automatica
    st.divider()
    st.subheader("🤖 Generazione Automatica")
    
    if st.button("🎯 Genera Programma Automatico", type="primary", use_container_width=True):
        if df_cani.empty or df_volontari.empty or df_luoghi.empty:
            st.error("❌ Carica prima i dati da Google Sheets (Cani, Volontari, Luoghi)")
        else:
            # Inizializzo programma
            programma = []
            
            # Aggiungo briefing se abilitato
            if briefing_enabled:
                programma.append({
                    "Orario": ora_briefing.strftime("%H:%M"),
                    "Cane": "TUTTI",
                    "Volontario": "TUTTI",
                    "Luogo": "Sala Briefing",
                    "Note": "Briefing giornaliero - Tutti i volontari"
                })
            
            # Filtro i luoghi per automatico
            if 'automatico' in df_luoghi.columns:
                luoghi_auto = df_luoghi[df_luoghi['automatico'].str.lower() == 'sì']['nome'].tolist()
            else:
                luoghi_auto = df_luoghi['nome'].tolist()
            
            # Preparo liste
            lista_cani = df_cani['nome'].tolist()
            lista_volontari = df_volontari['nome'].tolist()[:num_volontari]
            
            # Creo gli slot orari
            current_time = datetime.combine(data_t, ora_inizio)
            end_time = datetime.combine(data_t, ora_fine)
            
            idx_volontario = 0
            idx_luogo = 0
            
            while current_time < end_time:
                ora_str = current_time.strftime("%H:%M")
                
                # Aggiungo pasti se abilitato e siamo nell'orario giusto
                if pasti_enabled and current_time.time() == ora_pasti:
                    programma.append({
                        "Orario": ora_str,
                        "Cane": "TUTTI",
                        "Volontario": "TUTTI",
                        "Luogo": "Area Pasti",
                        "Note": "Pausa pasti - Tutti i volontari"
                    })
                    current_time += timedelta(minutes=durata)
                    continue
                
                # Per ogni slot, assegno tutti i volontari disponibili
                for _ in range(num_volontari):
                    if idx_volontario >= len(lista_volontari):
                        idx_volontario = 0
                    
                    volontario = lista_volontari[idx_volontario]
                    
                    # Scelgo un cane (rotazione semplice)
                    if lista_cani:
                        cane = lista_cani[idx_volontario % len(lista_cani)]
                        
                        # Verifico reattività e adiacenza
                        tentativi_campo = 0
                        campo_trovato = False
                        
                        while tentativi_campo < len(luoghi_auto):
                            campo = luoghi_auto[idx_luogo % len(luoghi_auto)]
                            
                            # Controllo validità campo
                            if campo_valido_per_reattivita(cane, campo, programma, ora_str, df_cani, df_luoghi):
                                # Campo valido, assegno il volontario più esperto
                                volontario_esperto = assegna_volontario_esperto(cane, lista_volontari, df_storico)
                                
                                programma.append({
                                    "Orario": ora_str,
                                    "Cane": cane,
                                    "Volontario": volontario_esperto,
                                    "Luogo": campo,
                                    "Note": ""
                                })
                                campo_trovato = True
                                idx_luogo += 1
                                break
                            else:
                                # Campo non valido, provo il successivo
                                idx_luogo += 1
                                tentativi_campo += 1
                        
                        if not campo_trovato:
                            # Nessun campo valido trovato, assegno comunque con nota di warning
                            st.warning(f"⚠️ Attenzione: nessun campo valido trovato per {cane} alle {ora_str} (reattività)")
                            programma.append({
                                "Orario": ora_str,
                                "Cane": cane,
                                "Volontario": volontario,
                                "Luogo": luoghi_auto[idx_luogo % len(luoghi_auto)],
                                "Note": "⚠️ Verifica reattività"
                            })
                            idx_luogo += 1
                    
                    idx_volontario += 1
                
                current_time += timedelta(minutes=durata)
            
            st.session_state.programma = programma
            st.success(f"✅ Programma generato: {len(programma)} turni creati!")
            st.rerun()
    
    # Mostra programma corrente
    if st.session_state.programma:
        st.divider()
        st.subheader("📋 Programma Corrente")
        
        df_programma = pd.DataFrame(st.session_state.programma)
        st.dataframe(df_programma, use_container_width=True, hide_index=True)
        
        # Modifica manuale
        st.divider()
        st.subheader("✏️ Aggiungi/Modifica Turno Manuale")
        
        col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns(5)
        
        with col_m1:
            ora_manuale = st.time_input("Orario", key="ora_man")
        
        with col_m2:
            cane_manuale = st.selectbox("Cane", [""] + df_cani['nome'].tolist(), key="cane_man")
        
        with col_m3:
            volontario_manuale = st.selectbox("Volontario", [""] + df_volontari['nome'].tolist(), key="vol_man")
        
        with col_m4:
            luogo_manuale = st.selectbox("Luogo", [""] + df_luoghi['nome'].tolist(), key="luogo_man")
        
        with col_m5:
            note_manuale = st.text_input("Note", key="note_man")
        
        col_add, col_clear = st.columns(2)
        
        with col_add:
            if st.button("➕ Aggiungi Turno", use_container_width=True):
                if cane_manuale and volontario_manuale and luogo_manuale:
                    nuovo_turno = {
                        "Orario": ora_manuale.strftime("%H:%M"),
                        "Cane": cane_manuale,
                        "Volontario": volontario_manuale,
                        "Luogo": luogo_manuale,
                        "Note": note_manuale
                    }
                    st.session_state.programma.append(nuovo_turno)
                    st.success("✅ Turno aggiunto!")
                    st.rerun()
                else:
                    st.error("❌ Compila tutti i campi obbligatori")
        
        with col_clear:
            if st.button("🗑️ Svuota Programma", use_container_width=True):
                st.session_state.programma = []
                st.success("✅ Programma svuotato!")
                st.rerun()

# ============================================================================
# TAB 2: GESTIONE PROGRAMMI SALVATI (NOVITÀ)
# ============================================================================
with tab_gestione:
    st.header("💾 Gestione Programmi Salvati")
    
    # Sottotab per organizzare le funzionalità
    subtab_salva, subtab_carica, subtab_lista, subtab_backup = st.tabs([
        "💾 Salva Programma",
        "📂 Carica Programma",
        "📑 Lista Programmi",
        "🔄 Backup Storico"
    ])
    
    # --- SALVA PROGRAMMA ---
    with subtab_salva:
        st.subheader("💾 Salva il Programma Corrente")
        
        if st.session_state.programma:
            st.info(f"📊 Programma corrente: **{len(st.session_state.programma)}** turni")
            
            # Preview del programma
            with st.expander("👁️ Visualizza Anteprima Programma"):
                df_preview = pd.DataFrame(st.session_state.programma)
                st.dataframe(df_preview, use_container_width=True, hide_index=True)
            
            st.divider()
            
            # Form di salvataggio
            with st.form("form_salva_programma"):
                col_s1, col_s2 = st.columns(2)
                
                with col_s1:
                    data_salvataggio = st.date_input("📅 Data del programma", value=data_t)
                
                with col_s2:
                    nome_programma = st.text_input("📝 Nome descrittivo (opzionale)", 
                                                   placeholder=f"Programma {data_salvataggio.strftime('%d/%m/%Y')}")
                
                note_programma = st.text_area("📋 Note (opzionale)", 
                                             placeholder="Aggiungi eventuali note o annotazioni...")
                
                col_sub1, col_sub2 = st.columns(2)
                
                with col_sub1:
                    salva_submitted = st.form_submit_button("💾 Salva Programma", 
                                                           type="primary", 
                                                           use_container_width=True)
                
                with col_sub2:
                    # Export Excel
                    export_excel = st.form_submit_button("📊 Esporta Excel", 
                                                        use_container_width=True)
                
                # Export PDF sotto
                export_pdf = st.form_submit_button("📄 Esporta PDF", 
                                                  use_container_width=True)
                
                if salva_submitted:
                    data_str = data_salvataggio.strftime('%Y-%m-%d')
                    if salva_programma(data_str, st.session_state.programma, nome_programma, note_programma):
                        st.success(f"✅ Programma salvato con successo per il {data_salvataggio.strftime('%d/%m/%Y')}!")
                        st.info("💡 Puoi ora chiudere l'app o modificare il codice senza perdere questo programma.")
                    else:
                        st.error("❌ Errore nel salvataggio del programma")
                
                if export_excel:
                    data_str = data_salvataggio.strftime('%Y-%m-%d')
                    excel_buffer = export_programma_excel(st.session_state.programma, data_str)
                    st.download_button(
                        label="⬇️ Scarica Excel",
                        data=excel_buffer,
                        file_name=f"Programma_{data_str}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                if export_pdf:
                    data_str = data_salvataggio.strftime('%Y-%m-%d')
                    pdf_buffer = export_programma_pdf(st.session_state.programma, data_str)
                    st.download_button(
                        label="⬇️ Scarica PDF",
                        data=pdf_buffer,
                        file_name=f"Programma_{data_str}.pdf",
                        mime="application/pdf"
                    )
        else:
            st.warning("⚠️ Nessun programma da salvare. Crea prima un programma nella tab 'Crea Programma'.")
    
    # --- CARICA PROGRAMMA ---
    with subtab_carica:
        st.subheader("📂 Carica un Programma Salvato")
        
        # Lista dei programmi disponibili
        df_programmi = lista_programmi_salvati()
        
        if not df_programmi.empty:
            st.write("**Programmi disponibili:**")
            
            # Formatto la visualizzazione
            df_display = df_programmi.copy()
            df_display['data'] = pd.to_datetime(df_display['data']).dt.strftime('%d/%m/%Y')
            df_display['creato_il'] = pd.to_datetime(df_display['creato_il']).dt.strftime('%d/%m/%Y %H:%M')
            
            # Selezione del programma
            programma_selezionato = st.selectbox(
                "Seleziona un programma",
                options=df_programmi['data'].tolist(),
                format_func=lambda x: f"{pd.to_datetime(x).strftime('%d/%m/%Y')} - {df_programmi[df_programmi['data']==x]['nome'].values[0] or 'Senza nome'}"
            )
            
            if programma_selezionato:
                programma_dati = carica_programma(programma_selezionato)
                
                if programma_dati:
                    # Mostra dettagli
                    col_d1, col_d2, col_d3 = st.columns(3)
                    
                    with col_d1:
                        st.metric("📅 Data", pd.to_datetime(programma_dati['data']).strftime('%d/%m/%Y'))
                    
                    with col_d2:
                        st.metric("📝 Nome", programma_dati['nome'] or "Senza nome")
                    
                    with col_d3:
                        st.metric("🕐 Creato il", pd.to_datetime(programma_dati['creato_il']).strftime('%d/%m/%Y %H:%M'))
                    
                    if programma_dati['note']:
                        st.info(f"📋 Note: {programma_dati['note']}")
                    
                    # Preview turni
                    with st.expander("👁️ Anteprima Turni", expanded=True):
                        df_turni = pd.DataFrame(programma_dati['turni'])
                        st.dataframe(df_turni, use_container_width=True, hide_index=True)
                        st.metric("📊 Totale turni", len(programma_dati['turni']))
                    
                    st.divider()
                    
                    # Azioni
                    col_a1, col_a2, col_a3 = st.columns(3)
                    
                    with col_a1:
                        if st.button("📥 Carica in Memoria", type="primary", use_container_width=True):
                            st.session_state.programma = programma_dati['turni']
                            st.success("✅ Programma caricato! Vai alla tab 'Crea Programma' per visualizzarlo.")
                            st.rerun()
                    
                    with col_a2:
                        excel_buffer = export_programma_excel(programma_dati['turni'], programma_selezionato)
                        st.download_button(
                            label="📊 Esporta Excel",
                            data=excel_buffer,
                            file_name=f"Programma_{programma_selezionato}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    
                    with col_a3:
                        pdf_buffer = export_programma_pdf(programma_dati['turni'], programma_selezionato)
                        st.download_button(
                            label="📄 Esporta PDF",
                            data=pdf_buffer,
                            file_name=f"Programma_{programma_selezionato}.pdf",
                            mime="application/pdf",
                            use_container_width=True
                        )
        else:
            st.info("📭 Nessun programma salvato. Salva prima un programma nella tab 'Salva Programma'.")
    
    # --- LISTA E GESTIONE ---
    with subtab_lista:
        st.subheader("📑 Tutti i Programmi Salvati")
        
        df_programmi = lista_programmi_salvati()
        
        if not df_programmi.empty:
            # Formatto per la visualizzazione
            df_display = df_programmi.copy()
            df_display['data'] = pd.to_datetime(df_display['data']).dt.strftime('%d/%m/%Y')
            df_display['creato_il'] = pd.to_datetime(df_display['creato_il']).dt.strftime('%d/%m/%Y %H:%M')
            df_display = df_display.drop('dim_dati', axis=1)
            
            st.dataframe(
                df_display,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "data": st.column_config.TextColumn("📅 Data", width="medium"),
                    "nome": st.column_config.TextColumn("📝 Nome", width="large"),
                    "creato_il": st.column_config.TextColumn("🕐 Creato il", width="medium"),
                    "note": st.column_config.TextColumn("📋 Note", width="large"),
                }
            )
            
            st.divider()
            st.subheader("🗑️ Cancellazione Programmi")
            
            col_del1, col_del2 = st.columns([3, 1])
            
            with col_del1:
                data_da_eliminare = st.selectbox(
                    "Seleziona programma da eliminare",
                    ["---"] + df_programmi['data'].tolist(),
                    format_func=lambda x: "---" if x == "---" else f"{pd.to_datetime(x).strftime('%d/%m/%Y')} - {df_programmi[df_programmi['data']==x]['nome'].values[0] or 'Senza nome'}"
                )
            
            with col_del2:
                if st.button("❌ Elimina", use_container_width=True, disabled=(data_da_eliminare == "---")):
                    if data_da_eliminare != "---":
                        if st.session_state.get('confirm_delete_prog') == data_da_eliminare:
                            elimina_programma(data_da_eliminare)
                            st.success(f"✅ Programma del {pd.to_datetime(data_da_eliminare).strftime('%d/%m/%Y')} eliminato!")
                            st.session_state.confirm_delete_prog = None
                            st.rerun()
                        else:
                            st.session_state.confirm_delete_prog = data_da_eliminare
                            st.warning("⚠️ Sei sicuro? Clicca di nuovo per confermare.")
        else:
            st.info("📭 Nessun programma salvato nel database.")
    
    # --- BACKUP STORICO ---
    with subtab_backup:
        st.subheader("🔄 Gestione Backup Storico")
        
        st.write("""
        **Protezione dello Storico**
        
        Lo storico dei turni è prezioso perché contiene l'esperienza dei volontari con ciascun cane.
        Qui puoi:
        - Creare backup manuali dello storico
        - Visualizzare i backup esistenti
        - Ripristinare lo storico da un backup precedente
        """)
        
        st.divider()
        
        # Crea backup
        col_b1, col_b2 = st.columns([3, 1])
        
        with col_b1:
            st.write("**Crea un nuovo backup dello storico**")
        
        with col_b2:
            if st.button("💾 Crea Backup", type="primary", use_container_width=True):
                count = backup_storico()
                st.success(f"✅ Backup #{count} creato con successo!")
                st.rerun()
        
        st.divider()
        
        # Lista backup
        st.write("**Backup disponibili**")
        
        conn = sqlite3.connect('canile.db')
        df_backups = pd.read_sql_query(
            """SELECT backup_timestamp, COUNT(*) as num_record 
               FROM storico_backup 
               GROUP BY backup_timestamp 
               ORDER BY backup_timestamp DESC""", 
            conn
        )
        conn.close()
        
        if not df_backups.empty:
            df_backups['backup_timestamp'] = pd.to_datetime(df_backups['backup_timestamp']).dt.strftime('%d/%m/%Y %H:%M:%S')
            
            st.dataframe(
                df_backups,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "backup_timestamp": st.column_config.TextColumn("🕐 Data/Ora Backup", width="medium"),
                    "num_record": st.column_config.NumberColumn("📊 Record", width="small"),
                }
            )
            
            # Ripristino
            st.divider()
            st.write("**Ripristina storico da backup**")
            
            st.warning("⚠️ ATTENZIONE: Il ripristino sovrascriverà lo storico attuale!")
            
            col_r1, col_r2 = st.columns([3, 1])
            
            with col_r1:
                backup_selezionato = st.selectbox(
                    "Seleziona backup da ripristinare",
                    ["---"] + df_backups['backup_timestamp'].tolist()
                )
            
            with col_r2:
                if st.button("🔄 Ripristina", use_container_width=True, disabled=(backup_selezionato == "---")):
                    if backup_selezionato != "---":
                        if st.session_state.get('confirm_restore') == backup_selezionato:
                            # Converto il timestamp nel formato corretto
                            backup_ts = pd.to_datetime(backup_selezionato, format='%d/%m/%Y %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
                            ripristina_storico_da_backup(backup_ts)
                            st.success(f"✅ Storico ripristinato dal backup del {backup_selezionato}!")
                            st.session_state.confirm_restore = None
                            st.rerun()
                        else:
                            st.session_state.confirm_restore = backup_selezionato
                            st.error("⚠️ CONFERMA: Questo sovrascriverà lo storico corrente. Clicca di nuovo per confermare.")
        else:
            st.info("📭 Nessun backup disponibile. Crea il primo backup!")

# ============================================================================
# TAB 3: ANAGRAFICA CANI (Codice esistente)
# ============================================================================
with tab_anagrafica:
    st.header("📋 Gestione Anagrafica Cani")
    
    subtab_carica_pdf, subtab_visualizza, subtab_modifica = st.tabs([
        "📄 Carica da PDF",
        "👁️ Visualizza Anagrafica",
        "✏️ Modifica Manuale"
    ])
    
    with subtab_carica_pdf:
        st.write("Carica un PDF contenente le informazioni del cane per estrarre automaticamente i campi.")
        
        uploaded_pdf = st.file_uploader("Scegli un file PDF", type=['pdf'], key="pdf_upload")
        
        if uploaded_pdf is not None:
            # Leggo il PDF
            pdf_reader = PyPDF2.PdfReader(uploaded_pdf)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text()
            
            # Mostro preview del testo estratto
            with st.expander("📖 Testo estratto dal PDF"):
                st.text_area("Contenuto", text, height=200, disabled=True)
            
            # Parsing
            st.divider()
            st.subheader("🔍 Dati Estratti")
            
            dati_estratti = parse_pdf_content(text)
            
            # Cerco il nome del cane nel testo (opzionale, migliorabile)
            nome_cane_trovato = ""
            # Potresti aggiungere qui logica per estrarre il nome dal PDF
            
            col_nome, col_conferma = st.columns([3, 1])
            
            with col_nome:
                nome_cane = st.text_input("📝 Nome del Cane", value=nome_cane_trovato, 
                                         placeholder="Inserisci il nome del cane")
            
            # Mostro i campi estratti in modo editabile
            st.write("**Verifica e modifica i dati estratti:**")
            
            cibo_edit = st.text_area("🍖 CIBO", value=dati_estratti.get('CIBO', 'N/D'), height=100)
            guinz_edit = st.text_area("🎯 GUINZAGLIERIA", value=dati_estratti.get('GUINZAGLIERIA', 'N/D'), height=100)
            strum_edit = st.text_area("🛠️ STRUMENTI", value=dati_estratti.get('STRUMENTI', 'N/D'), height=100)
            attiv_edit = st.text_area("🎾 ATTIVITÀ", value=dati_estratti.get('ATTIVITÀ', 'N/D'), height=100)
            note_edit = st.text_area("📝 NOTE", value=dati_estratti.get('NOTE', 'N/D'), height=100)
            tempo_edit = st.text_area("⏱️ TEMPO", value=dati_estratti.get('TEMPO', 'N/D'), height=100)
            
            with col_conferma:
                st.write("")  # Spacing
                st.write("")
                if st.button("💾 Salva in Anagrafica", type="primary", use_container_width=True):
                    if nome_cane.strip():
                        dati_da_salvare = {
                            'CIBO': cibo_edit,
                            'GUINZAGLIERIA': guinz_edit,
                            'STRUMENTI': strum_edit,
                            'ATTIVITÀ': attiv_edit,
                            'NOTE': note_edit,
                            'TEMPO': tempo_edit
                        }
                        
                        if salva_info_cane(nome_cane, dati_da_salvare):
                            st.success(f"✅ Informazioni salvate per {nome_cane.capitalize()}!")
                    else:
                        st.error("❌ Inserisci il nome del cane")
    
    with subtab_visualizza:
        st.write("Visualizza le schede complete dei cani presenti in anagrafica.")
        
        # Carico l'anagrafica
        conn = sqlite3.connect('canile.db')
        df_anagrafica = pd.read_sql_query("SELECT nome FROM anagrafica_cani ORDER BY nome", conn)
        conn.close()
        
        if not df_anagrafica.empty:
            cane_selezionato = st.selectbox("Seleziona un cane", df_anagrafica['nome'].tolist())
            
            if cane_selezionato:
                info_cane = get_info_cane(cane_selezionato)
                
                st.divider()
                st.subheader(f"🐕 Scheda di {cane_selezionato.upper()}")
                
                col_v1, col_v2 = st.columns(2)
                
                with col_v1:
                    st.write("**🍖 CIBO**")
                    st.info(info_cane['CIBO'])
                    
                    st.write("**🎯 GUINZAGLIERIA**")
                    st.info(info_cane['GUINZAGLIERIA'])
                    
                    st.write("**🛠️ STRUMENTI**")
                    st.info(info_cane['STRUMENTI'])
                
                with col_v2:
                    st.write("**🎾 ATTIVITÀ**")
                    st.info(info_cane['ATTIVITÀ'])
                    
                    st.write("**📝 NOTE**")
                    st.info(info_cane['NOTE'])
                    
                    st.write("**⏱️ TEMPO**")
                    st.info(info_cane['TEMPO'])
        else:
            st.info("📭 Nessun cane presente in anagrafica. Carica dei PDF per iniziare!")
    
    with subtab_modifica:
        st.write("Modifica manualmente le informazioni di un cane o creane uno nuovo.")
        
        # Carico l'anagrafica
        conn = sqlite3.connect('canile.db')
        df_anagrafica = pd.read_sql_query("SELECT nome FROM anagrafica_cani ORDER BY nome", conn)
        conn.close()
        
        col_tipo, col_nome_mod = st.columns([1, 2])
        
        with col_tipo:
            tipo_azione = st.radio("Azione", ["Modifica esistente", "Nuovo cane"])
        
        with col_nome_mod:
            if tipo_azione == "Modifica esistente":
                if not df_anagrafica.empty:
                    nome_da_modificare = st.selectbox("Seleziona cane", df_anagrafica['nome'].tolist())
                    info_corrente = get_info_cane(nome_da_modificare)
                else:
                    st.warning("Nessun cane in anagrafica")
                    nome_da_modificare = None
                    info_corrente = {k: "N/D" for k in ['CIBO', 'GUINZAGLIERIA', 'STRUMENTI', 'ATTIVITÀ', 'NOTE', 'TEMPO']}
            else:
                nome_da_modificare = st.text_input("Nome del nuovo cane")
                info_corrente = {k: "N/D" for k in ['CIBO', 'GUINZAGLIERIA', 'STRUMENTI', 'ATTIVITÀ', 'NOTE', 'TEMPO']}
        
        st.divider()
        
        if (tipo_azione == "Modifica esistente" and nome_da_modificare) or (tipo_azione == "Nuovo cane" and nome_da_modificare):
            st.subheader(f"✏️ Modifica: {nome_da_modificare}")
            
            cibo_mod = st.text_area("🍖 CIBO", value=info_corrente.get('CIBO', 'N/D'), height=100, key="cibo_mod")
            guinz_mod = st.text_area("🎯 GUINZAGLIERIA", value=info_corrente.get('GUINZAGLIERIA', 'N/D'), height=100, key="guinz_mod")
            strum_mod = st.text_area("🛠️ STRUMENTI", value=info_corrente.get('STRUMENTI', 'N/D'), height=100, key="strum_mod")
            attiv_mod = st.text_area("🎾 ATTIVITÀ", value=info_corrente.get('ATTIVITÀ', 'N/D'), height=100, key="attiv_mod")
            note_mod = st.text_area("📝 NOTE", value=info_corrente.get('NOTE', 'N/D'), height=100, key="note_mod")
            tempo_mod = st.text_area("⏱️ TEMPO", value=info_corrente.get('TEMPO', 'N/D'), height=100, key="tempo_mod")
            
            if st.button("💾 Salva Modifiche", type="primary", use_container_width=True):
                dati_modificati = {
                    'CIBO': cibo_mod,
                    'GUINZAGLIERIA': guinz_mod,
                    'STRUMENTI': strum_mod,
                    'ATTIVITÀ': attiv_mod,
                    'NOTE': note_mod,
                    'TEMPO': tempo_mod
                }
                
                if salva_info_cane(nome_da_modificare, dati_modificati):
                    st.success(f"✅ Modifiche salvate per {nome_da_modificare}!")
                    st.rerun()

# ============================================================================
# TAB 4: STORICO & STATISTICHE (Codice esistente)
# ============================================================================
with tab_storico:
    st.header("📊 Storico Turni e Statistiche")
    
    subtab_visualizza, subtab_stats = st.tabs(["📋 Visualizza & Modifica", "📈 Statistiche"])
    
    with subtab_visualizza:
        st.write("### 📋 Storico Completo dei Turni")
        
        conn = sqlite3.connect('canile.db')
        df_storico_completo = pd.read_sql_query("SELECT rowid, * FROM storico ORDER BY data DESC, inizio", conn)
        
        if not df_storico_completo.empty:
            # Converto la data in formato datetime per il filtro
            df_storico_completo['data'] = pd.to_datetime(df_storico_completo['data'])
            
            # Filtri
            st.write("#### 🔍 Filtri")
            col_f1, col_f2, col_f3 = st.columns(3)
            
            with col_f1:
                date_uniche = sorted(df_storico_completo['data'].dt.strftime('%Y-%m-%d').unique(), reverse=True)
                filtro_data = st.selectbox("Filtra per data", ["Tutte"] + date_uniche, key="filtro_data_storico")
            
            with col_f2:
                cani_unici = sorted(df_storico_completo['cane'].unique())
                filtro_cane = st.selectbox("Filtra per cane", ["Tutti"] + cani_unici, key="filtro_cane_storico")
            
            with col_f3:
                volontari_unici = sorted(df_storico_completo['volontario'].unique())
                filtro_volontario = st.selectbox("Filtra per volontario", ["Tutti"] + volontari_unici, key="filtro_vol_storico")
            
            # Applico filtri
            df_filtrato = df_storico_completo.copy()
            
            if filtro_data != "Tutte":
                df_filtrato = df_filtrato[df_filtrato['data'].dt.strftime('%Y-%m-%d') == filtro_data]
            
            if filtro_cane != "Tutti":
                df_filtrato = df_filtrato[df_filtrato['cane'] == filtro_cane]
            
            if filtro_volontario != "Tutti":
                df_filtrato = df_filtrato[df_filtrato['volontario'] == filtro_volontario]
            
            st.divider()
            
            if not df_filtrato.empty:
                st.write(f"**{len(df_filtrato)} turni trovati**")
                
                # Data editor per modifiche
                df_edited = st.data_editor(
                    df_filtrato[['rowid', 'data', 'inizio', 'cane', 'volontario', 'luogo']],
                    use_container_width=True,
                    hide_index=True,
                    num_rows="dynamic",
                    column_config={
                        "rowid": st.column_config.NumberColumn("ID", disabled=True),
                        "data": st.column_config.DateColumn("Data", format="DD/MM/YYYY"),
                        "inizio": st.column_config.TextColumn("Orario"),
                        "cane": st.column_config.TextColumn("Cane"),
                        "volontario": st.column_config.TextColumn("Volontario"),
                        "luogo": st.column_config.TextColumn("Luogo"),
                    }
                )
                
                # Pulsanti di azione
                col_save, col_del = st.columns(2)
                
                with col_save:
                    if st.button("💾 Salva Modifiche", use_container_width=True):
                        try:
                            # Aggiorno tutti i record modificati
                            for _, row in df_edited.iterrows():
                                # Converto la data da datetime a stringa formato YYYY-MM-DD
                                data_str = row['data'].strftime('%Y-%m-%d') if pd.notna(row['data']) else row['data']
                                conn.execute(
                                    "UPDATE storico SET data=?, inizio=?, cane=?, volontario=?, luogo=? WHERE rowid=?",
                                    (data_str, row['inizio'], row['cane'], row['volontario'], row['luogo'], row['rowid'])
                                )
                            conn.commit()
                            st.success("✅ Modifiche salvate con successo!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Errore nel salvataggio: {str(e)}")
                
                with col_del:
                    if st.button("🗑️ Elimina Turni Selezionati", use_container_width=True):
                        st.warning("⚠️ Funzione in sviluppo: usa il data editor per eliminare righe (modalità 'dynamic')")
                
                st.divider()
                
                # Cancellazione rapida per data
                st.write("#### 🗑️ Cancellazione Rapida")
                col_del1, col_del2 = st.columns([2, 1])
                
                with col_del1:
                    data_da_cancellare = st.selectbox("Seleziona data da cancellare completamente", 
                                                      ["---"] + date_uniche,
                                                      key="del_data")
                
                with col_del2:
                    if st.button("❌ Cancella Giornata", use_container_width=True, disabled=(data_da_cancellare == "---")):
                        if data_da_cancellare != "---":
                            count = conn.execute("SELECT COUNT(*) FROM storico WHERE data=?", (data_da_cancellare,)).fetchone()[0]
                            if st.session_state.get('confirm_delete') == data_da_cancellare:
                                conn.execute("DELETE FROM storico WHERE data=?", (data_da_cancellare,))
                                conn.commit()
                                st.success(f"✅ Cancellati {count} turni del {data_da_cancellare}")
                                st.session_state.confirm_delete = None
                                st.rerun()
                            else:
                                st.session_state.confirm_delete = data_da_cancellare
                                st.warning(f"⚠️ Vuoi davvero cancellare {count} turni? Clicca di nuovo per confermare.")
            else:
                st.info("Nessun turno trovato con i filtri selezionati.")
        else:
            st.info("📭 Nessun dato nello storico. Salva alcuni turni per iniziare!")
    
    with subtab_stats:
        st.write("### 📈 Statistiche Esperienza")
        
        df_storico = pd.read_sql_query("SELECT * FROM storico", conn)
        
        if not df_storico.empty:
            # Statistiche per cane
            st.write("#### 🐕 Esperienza per Cane")
            
            cani_disponibili = sorted(df_storico['cane'].unique())
            cane_selezionato = st.selectbox("Seleziona un cane", cani_disponibili, key="stats_cane")
            
            if cane_selezionato:
                # Statistiche del cane selezionato
                df_cane = df_storico[df_storico['cane'] == cane_selezionato]
                
                # Raggruppo per volontario
                stats_volontari = df_cane.groupby('volontario').agg({
                    'data': 'count'
                }).reset_index()
                stats_volontari.columns = ['Volontario', 'Turni Totali']
                stats_volontari = stats_volontari.sort_values('Turni Totali', ascending=False)
                
                col_stat1, col_stat2 = st.columns(2)
                
                with col_stat1:
                    st.metric("📊 Turni Totali con questo cane", len(df_cane))
                    st.metric("👥 Volontari Diversi", len(stats_volontari))
                
                with col_stat2:
                    if len(stats_volontari) > 0:
                        st.metric("🥇 Volontario più esperto", 
                                 stats_volontari.iloc[0]['Volontario'],
                                 f"{stats_volontari.iloc[0]['Turni Totali']} turni")
                
                st.divider()
                st.write("**Classifica Esperienza:**")
                
                # Aggiungo una colonna con la percentuale
                stats_volontari['Percentuale'] = (stats_volontari['Turni Totali'] / len(df_cane) * 100).round(1)
                
                st.dataframe(
                    stats_volontari,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Volontario": st.column_config.TextColumn("👤 Volontario", width="medium"),
                        "Turni Totali": st.column_config.NumberColumn("📊 Turni", width="small"),
                        "Percentuale": st.column_config.NumberColumn("📈 %", format="%.1f%%", width="small"),
                    }
                )
                
                # Grafico a barre
                st.bar_chart(stats_volontari.set_index('Volontario')['Turni Totali'])
            
            st.divider()
            
            # Statistiche generali
            st.write("#### 📊 Statistiche Generali")
            
            col_g1, col_g2, col_g3 = st.columns(3)
            
            with col_g1:
                st.metric("🐕 Cani Totali", df_storico['cane'].nunique())
                st.metric("👥 Volontari Totali", df_storico['volontario'].nunique())
            
            with col_g2:
                st.metric("📅 Giorni con Turni", df_storico['data'].nunique())
                st.metric("📍 Luoghi Utilizzati", df_storico['luogo'].nunique())
            
            with col_g3:
                st.metric("✅ Turni Totali", len(df_storico))
                media_turni_giorno = len(df_storico) / max(df_storico['data'].nunique(), 1)
                st.metric("📊 Media Turni/Giorno", f"{media_turni_giorno:.1f}")
            
            # Top volontari
            st.write("#### 🏆 Top 10 Volontari più Attivi")
            top_volontari = df_storico.groupby('volontario').size().reset_index(name='Turni')
            top_volontari = top_volontari.sort_values('Turni', ascending=False).head(10)
            
            st.dataframe(
                top_volontari.reset_index(drop=True),
                use_container_width=True,
                hide_index=True,
                column_config={
                    "volontario": st.column_config.TextColumn("👤 Volontario", width="large"),
                    "Turni": st.column_config.NumberColumn("📊 Turni Totali", width="medium"),
                }
            )
            
        else:
            st.info("📭 Nessun dato disponibile per le statistiche. Salva alcuni turni per iniziare!")
    
    conn.close()

# --- SEZIONE SALVATAGGIO IN STORICO ---
st.divider()
st.subheader("💾 Salvataggio Giornata in Storico")

if st.session_state.programma:
    st.info(f"📊 Turni programmati: **{len(st.session_state.programma)}** (verranno salvati solo i turni con cani specifici)")
    
    col_salva1, col_salva2 = st.columns([3, 1])
    
    with col_salva1:
        st.write("Una volta completata la giornata, salva i turni nello storico per migliorare l'assegnazione automatica futura.")
    
    with col_salva2:
        if st.button("✅ Conferma e Salva in Storico", type="primary", use_container_width=True):
            record_salvati = salva_turni_in_storico(st.session_state.programma, data_t)
            if record_salvati > 0:
                st.success(f"✅ Salvati {record_salvati} turni nello storico del {data_t.strftime('%d/%m/%Y')}!")
                st.info("💡 L'algoritmo di assegnazione automatica ora terrà conto di questi turni per dare priorità ai volontari più esperti con ogni cane.")
                # Opzionalmente: svuoto il programma dopo il salvataggio
                # st.session_state.programma = []
                # st.rerun()
            else:
                st.warning("⚠️ Nessun turno valido da salvare (solo turni speciali o senza cane).")
else:
    st.info("📝 Crea prima un programma giornaliero per poterlo salvare nello storico.")
